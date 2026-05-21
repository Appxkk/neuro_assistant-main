import time
import json
from pathlib import Path
from datetime import datetime

import httpx
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


BASE_URL = "http://127.0.0.1:8000"

REPORT_DIR = Path("tests/reports")
REPORT_DIR.mkdir(parents=True, exist_ok=True)


class TestResult:
    def __init__(
        self,
        number: int,
        name: str,
        check_description: str,
        module: str,
        method: str,
        endpoint: str,
        input_data: str,
        expected_result: str,
        actual_result: str,
        status: str,
        start_time: datetime,
        end_time: datetime,
        duration: float,
        comment: str = "",
    ):
        self.number = number
        self.name = name
        self.check_description = check_description
        self.module = module
        self.method = method
        self.endpoint = endpoint
        self.input_data = input_data
        self.expected_result = expected_result
        self.actual_result = actual_result
        self.status = status
        self.start_time = start_time
        self.end_time = end_time
        self.duration = duration
        self.comment = comment


def now_str(dt: datetime) -> str:
    return dt.strftime("%d.%m.%Y %H:%M:%S")


def to_json(value) -> str:
    if value is None:
        return ""

    try:
        return json.dumps(value, ensure_ascii=False, indent=2)
    except Exception:
        return str(value)


def make_check_description(
    module: str,
    method: str,
    endpoint: str,
    name: str,
    expected_field: str | None = None,
    expected_value=None,
) -> str:
    description = (
        f"Проверяется функция: «{name}». "
        f"Тест относится к модулю «{module}». "
        f"Выполняется {method.upper()}-запрос к endpoint «{endpoint}». "
        f"Система должна вернуть корректный ответ и подтвердить работоспособность проверяемой функции."
    )

    if expected_field:
        description += (
            f" Дополнительно проверяется значение поля «{expected_field}»: "
            f"ожидаемое значение — «{expected_value}»."
        )

    return description


def run_http_test(
    client: httpx.Client,
    number: int,
    name: str,
    module: str,
    method: str,
    endpoint: str,
    expected_status: int = 200,
    expected_field: str | None = None,
    expected_value=None,
    json_body: dict | None = None,
    expect_binary: bool = False,
) -> TestResult:
    start_time = datetime.now()
    started = time.perf_counter()

    input_data = to_json(json_body)
    expected_result = f"HTTP {expected_status}"

    if expected_field:
        expected_result += f", поле {expected_field} = {expected_value}"

    check_description = make_check_description(
        module=module,
        method=method,
        endpoint=endpoint,
        name=name,
        expected_field=expected_field,
        expected_value=expected_value,
    )

    try:
        if method.upper() == "GET":
            response = client.get(endpoint)
        elif method.upper() == "POST":
            response = client.post(endpoint, json=json_body)
        elif method.upper() == "DELETE":
            response = client.delete(endpoint)
        else:
            raise ValueError(f"Неподдерживаемый метод: {method}")

        end_time = datetime.now()
        duration = round(time.perf_counter() - started, 4)

        if response.status_code != expected_status:
            return TestResult(
                number=number,
                name=name,
                check_description=check_description,
                module=module,
                method=method,
                endpoint=endpoint,
                input_data=input_data,
                expected_result=expected_result,
                actual_result=f"HTTP {response.status_code}: {response.text[:500]}",
                status="Ошибка",
                start_time=start_time,
                end_time=end_time,
                duration=duration,
                comment="Код ответа не совпадает с ожидаемым",
            )

        if expect_binary:
            content_length = len(response.content)
            ok = content_length > 0

            return TestResult(
                number=number,
                name=name,
                check_description=check_description,
                module=module,
                method=method,
                endpoint=endpoint,
                input_data=input_data,
                expected_result="Получен бинарный файл",
                actual_result=f"Размер файла: {content_length} байт",
                status="Успешно" if ok else "Ошибка",
                start_time=start_time,
                end_time=end_time,
                duration=duration,
                comment="" if ok else "Файл пустой",
            )

        try:
            data = response.json()
        except Exception:
            data = {"raw_response": response.text}

        if expected_field:
            actual_value = data

            for part in expected_field.split("."):
                if isinstance(actual_value, dict):
                    actual_value = actual_value.get(part)
                else:
                    actual_value = None
                    break

            if actual_value != expected_value:
                return TestResult(
                    number=number,
                    name=name,
                    check_description=check_description,
                    module=module,
                    method=method,
                    endpoint=endpoint,
                    input_data=input_data,
                    expected_result=expected_result,
                    actual_result=f"{expected_field} = {actual_value}",
                    status="Ошибка",
                    start_time=start_time,
                    end_time=end_time,
                    duration=duration,
                    comment="Значение поля не совпадает с ожидаемым",
                )

        return TestResult(
            number=number,
            name=name,
            check_description=check_description,
            module=module,
            method=method,
            endpoint=endpoint,
            input_data=input_data,
            expected_result=expected_result,
            actual_result="Проверка выполнена успешно",
            status="Успешно",
            start_time=start_time,
            end_time=end_time,
            duration=duration,
            comment="",
        )

    except Exception as error:
        end_time = datetime.now()
        duration = round(time.perf_counter() - started, 4)

        return TestResult(
            number=number,
            name=name,
            check_description=check_description,
            module=module,
            method=method,
            endpoint=endpoint,
            input_data=input_data,
            expected_result=expected_result,
            actual_result=str(error),
            status="Ошибка",
            start_time=start_time,
            end_time=end_time,
            duration=duration,
            comment="Исключение во время выполнения теста",
        )


def build_test_cases():
    return [
        {
            "name": "Проверка доступности главной страницы",
            "module": "Web interface",
            "method": "GET",
            "endpoint": "/",
            "expected_status": 200,
        },
        {
            "name": "Загрузка dashboard",
            "module": "Dashboard API",
            "method": "GET",
            "endpoint": "/api/dashboard",
            "expected_status": 200,
        },
        {
            "name": "Загрузка списка заказов",
            "module": "Pages API",
            "method": "GET",
            "endpoint": "/api/pages/orders",
            "expected_status": 200,
        },
        {
            "name": "Загрузка списка задач",
            "module": "Pages API",
            "method": "GET",
            "endpoint": "/api/pages/tasks",
            "expected_status": 200,
        },
        {
            "name": "Загрузка отчётов",
            "module": "Pages API",
            "method": "GET",
            "endpoint": "/api/pages/reports",
            "expected_status": 200,
        },
        {
            "name": "Загрузка истории команд",
            "module": "Pages API",
            "method": "GET",
            "endpoint": "/api/pages/history",
            "expected_status": 200,
        },
        {
            "name": "Загрузка менеджеров для окна подтверждения",
            "module": "Actions API",
            "method": "GET",
            "endpoint": "/api/actions/managers",
            "expected_status": 200,
        },
        {
            "name": "Загрузка сотрудников для окна подтверждения",
            "module": "Actions API",
            "method": "GET",
            "endpoint": "/api/actions/employees",
            "expected_status": 200,
        },
        {
            "name": "Загрузка заказов для создания продажи",
            "module": "Actions API",
            "method": "GET",
            "endpoint": "/api/actions/orders",
            "expected_status": 200,
        },
        {
            "name": "Справочный чат: как добавить заказ",
            "module": "Help chat",
            "method": "POST",
            "endpoint": "/api/help/chat",
            "expected_status": 200,
            "json_body": {
                "question": "Как добавить заказ?"
            },
        },
        {
            "name": "Команда навигации: открыть настройки",
            "module": "Voice text API",
            "method": "POST",
            "endpoint": "/api/voice/text",
            "expected_status": 200,
            "expected_field": "intent",
            "expected_value": "open_page",
            "json_body": {
                "text": "Открой настройки"
            },
        },
        {
            "name": "Команда темы: включить тёмную тему",
            "module": "Voice text API",
            "method": "POST",
            "endpoint": "/api/voice/text",
            "expected_status": 200,
            "expected_field": "intent",
            "expected_value": "change_theme",
            "json_body": {
                "text": "Включи тёмную тему"
            },
        },
        {
            "name": "Команда графика",
            "module": "Voice text API",
            "method": "POST",
            "endpoint": "/api/voice/text",
            "expected_status": 200,
            "expected_field": "intent",
            "expected_value": "build_chart",
            "json_body": {
                "text": "Построй график"
            },
        },
        {
            "name": "Команда скачивания PDF",
            "module": "Voice text API",
            "method": "POST",
            "endpoint": "/api/voice/text",
            "expected_status": 200,
            "expected_field": "intent",
            "expected_value": "download_pdf",
            "json_body": {
                "text": "Скачай PDF"
            },
        },
        {
            "name": "Команда справки",
            "module": "Voice text API",
            "method": "POST",
            "endpoint": "/api/voice/text",
            "expected_status": 200,
            "expected_field": "intent",
            "expected_value": "help_question",
            "json_body": {
                "text": "Как добавить менеджера?"
            },
        },
        {
            "name": "Создание confirm_action для заказа",
            "module": "Pending action parser",
            "method": "POST",
            "endpoint": "/api/voice/text",
            "expected_status": 200,
            "expected_field": "parameters.action",
            "expected_value": "create_order",
            "json_body": {
                "text": "Добавь заказ клиента Тестовый клиент на сумму 1000"
            },
        },
        {
            "name": "Создание confirm_action для сотрудника",
            "module": "Pending action parser",
            "method": "POST",
            "endpoint": "/api/voice/text",
            "expected_status": 200,
            "expected_field": "parameters.action",
            "expected_value": "create_employee",
            "json_body": {
                "text": "Добавь сотрудника Тестов Тест должность Менеджер отдел Продажи"
            },
        },
        {
            "name": "Создание confirm_action для продажи",
            "module": "Pending action parser",
            "method": "POST",
            "endpoint": "/api/voice/text",
            "expected_status": 200,
            "expected_field": "parameters.action",
            "expected_value": "create_sale",
            "json_body": {
                "text": "Добавь продажу по заказу 1 товар Тестовый товар количество 1 на сумму 1000"
            },
        },
        {
            "name": "Создание confirm_action для задачи",
            "module": "Pending action parser",
            "method": "POST",
            "endpoint": "/api/voice/text",
            "expected_status": 200,
            "expected_field": "parameters.action",
            "expected_value": "create_task",
            "json_body": {
                "text": "Создай задачу тест описание Проверить систему исполнитель 1 приоритет высокий"
            },
        },
        {
            "name": "Создание confirm_action для Telegram-рассылки",
            "module": "Pending action parser",
            "method": "POST",
            "endpoint": "/api/voice/text",
            "expected_status": 200,
            "expected_field": "parameters.action",
            "expected_value": "send_mailing",
            "json_body": {
                "text": "Надо сделать рассылку"
            },
        },
        {
            "name": "Генерация PDF-файла",
            "module": "PDF API",
            "method": "POST",
            "endpoint": "/api/pdf/download",
            "expected_status": 200,
            "expect_binary": True,
            "json_body": {
                "recognized_text": "Тестовая команда",
                "intent": "test",
                "source": "test_runner",
                "result_text": "Тестовая генерация PDF",
                "parameters": {},
                "data": []
            },
        },
    ]


def create_excel_report(results: list[TestResult], global_start: datetime, global_end: datetime):
    report_path = REPORT_DIR / "neuro_assistant_test_report.xlsx"

    wb = Workbook()

    summary_ws = wb.active
    summary_ws.title = "Summary"

    detail_ws = wb.create_sheet("Test cases")

    success_count = sum(1 for item in results if item.status == "Успешно")
    error_count = sum(1 for item in results if item.status == "Ошибка")
    total_count = len(results)
    total_duration = round((global_end - global_start).total_seconds(), 4)

    summary_data = [
        ["Параметр", "Значение"],
        ["Название системы", "Нейропомощник для автоматизации бизнес-процессов"],
        ["Тип тестирования", "Автоматизированное функциональное тестирование API"],
        ["Дата и время начала", now_str(global_start)],
        ["Дата и время окончания", now_str(global_end)],
        ["Общее время тестирования, сек.", total_duration],
        ["Всего тестов", total_count],
        ["Успешно", success_count],
        ["Ошибок", error_count],
        ["Итоговый статус", "Успешно" if error_count == 0 else "Есть ошибки"],
    ]

    for row in summary_data:
        summary_ws.append(row)

    headers = [
        "№",
        "Название теста",
        "Что проверяется",
        "Модуль",
        "Метод",
        "Endpoint",
        "Входные данные",
        "Ожидаемый результат",
        "Фактический результат",
        "Статус",
        "Время начала",
        "Время окончания",
        "Длительность, сек.",
        "Комментарий",
    ]

    detail_ws.append(headers)

    for result in results:
        detail_ws.append([
            result.number,
            result.name,
            result.check_description,
            result.module,
            result.method,
            result.endpoint,
            result.input_data,
            result.expected_result,
            result.actual_result,
            result.status,
            now_str(result.start_time),
            now_str(result.end_time),
            result.duration,
            result.comment,
        ])

    style_workbook(wb)

    wb.save(report_path)

    return report_path


def style_workbook(wb: Workbook):
    header_fill = PatternFill("solid", fgColor="1F4E78")
    success_fill = PatternFill("solid", fgColor="D9EAD3")
    error_fill = PatternFill("solid", fgColor="F4CCCC")
    summary_fill = PatternFill("solid", fgColor="D9EAF7")

    white_font = Font(color="FFFFFF", bold=True)
    bold_font = Font(bold=True)
    border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    for ws in wb.worksheets:
        ws.freeze_panes = "A2"

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = white_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

        for row in ws.iter_rows():
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        for column_cells in ws.columns:
            max_length = 0
            column = column_cells[0].column

            for cell in column_cells:
                value = str(cell.value) if cell.value is not None else ""
                max_length = max(max_length, len(value))

            adjusted_width = min(max(max_length + 2, 14), 65)
            ws.column_dimensions[get_column_letter(column)].width = adjusted_width

    summary_ws = wb["Summary"]

    for row in summary_ws.iter_rows(min_row=2):
        row[0].font = bold_font
        row[0].fill = summary_fill

    detail_ws = wb["Test cases"]

    for row in detail_ws.iter_rows(min_row=2):
        status_cell = row[9]

        if status_cell.value == "Успешно":
            status_cell.fill = success_fill
            status_cell.font = Font(bold=True, color="38761D")

        if status_cell.value == "Ошибка":
            status_cell.fill = error_fill
            status_cell.font = Font(bold=True, color="990000")


def main():
    global_start = datetime.now()

    print("Запуск автоматизированного тестирования нейропомощника")
    print(f"Адрес тестируемой системы: {BASE_URL}")
    print(f"Время начала: {now_str(global_start)}")
    print("-" * 70)

    results = []
    test_cases = build_test_cases()

    with httpx.Client(base_url=BASE_URL, timeout=15) as client:
        for index, case in enumerate(test_cases, start=1):
            print(f"[{index}/{len(test_cases)}] {case['name']}...")

            result = run_http_test(
                client=client,
                number=index,
                name=case["name"],
                module=case["module"],
                method=case["method"],
                endpoint=case["endpoint"],
                expected_status=case.get("expected_status", 200),
                expected_field=case.get("expected_field"),
                expected_value=case.get("expected_value"),
                json_body=case.get("json_body"),
                expect_binary=case.get("expect_binary", False),
            )

            results.append(result)

            status_icon = "OK" if result.status == "Успешно" else "ERROR"
            print(f"    {status_icon} | {result.status} | {result.duration} сек.")

    global_end = datetime.now()

    report_path = create_excel_report(results, global_start, global_end)

    success_count = sum(1 for item in results if item.status == "Успешно")
    error_count = sum(1 for item in results if item.status == "Ошибка")
    total_duration = round((global_end - global_start).total_seconds(), 4)

    print("-" * 70)
    print("Тестирование завершено")
    print(f"Время окончания: {now_str(global_end)}")
    print(f"Общее время: {total_duration} сек.")
    print(f"Всего тестов: {len(results)}")
    print(f"Успешно: {success_count}")
    print(f"Ошибок: {error_count}")
    print(f"Отчёт сохранён: {report_path}")


if __name__ == "__main__":
    main()
